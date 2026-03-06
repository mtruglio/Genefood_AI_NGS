from openpyxl import load_workbook
import pandas as pd
from pprint import pprint
from .read_NGS_results import build_pandas_variant_db
from utils.errors import ValidationError as ValidationError

def check_missing_data(datadict, foglio, legend):
    # print("Check foglio", foglio)
    # print(datadict)
    lengths=[]
    for d in datadict:
        lengths.append(len(d))
    
    lengths_unique = set(lengths)
    if len(lengths_unique) != 1:
        real_length = max(lengths_unique)
        patients_missing_data = []
        for i in range(0, len(lengths)):
            if lengths[i]!=real_length:
                patients_missing_data.append(str(legend[i]))
        return "Errore: Mancano uno o piu' dati nel campo {0} sul foglio {1} ".format(','.join(set(patients_missing_data)),foglio)
    else:
        return
        
def check_wrong_data(paziente):
    # print("Check foglio", foglio)
    # print(datadict)

    return "Errore: Il codice di accettazione non e' corretto (deve essere 8 cifre). Paziente {0}".format(paziente)


def calc_weights_dict(df):
    weights_dict = {}
    for i, row in df.iterrows():
        gene = row["GENE"].strip()
        snp = row["SNP"].strip()
        genotype = row["GEN"].strip()
        weight = float(row["PESO"])
        if gene not in weights_dict:
            weights_dict[gene] = {}
            if '/' in gene:
                for g in gene.split('/'):
                    weights_dict[g] = {}
        if snp not in weights_dict[gene]:
            weights_dict[gene][snp] = {}
        if genotype not in weights_dict[gene][snp]:
            weights_dict[gene][snp][genotype] = weight
        else:
            print("error", gene, snp, genotype, "already present")
    
    #pprint(weights_dict)
    return weights_dict

def get_notes_dict(df):
    notes_dict ={}
    for i, row in df.iterrows():
        gene = row["GENE"].strip()
        snp = row["SNP"].strip()
        genesplit = row["GENE"].strip().split('/')
        snpsplit = row["SNP"].strip().split('/')
        
        genotype = row["GEN"].strip()
        note = row["NOTE"]
        if gene not in notes_dict:
            notes_dict[gene] = {}
        
        # Adding slashed genes and snps as they are
        if snp not in notes_dict[gene]:
                notes_dict[gene][snp] = {}
        snp_notes=[]
        # print("Note for gene {0} snp {1}: {2}".format(gene, snp, note))
        for n in note.split('+'):
            snp_notes.append(n.rstrip().lstrip())
        notes_dict[gene][snp] = snp_notes

        # Also adding each gene and snp in slashes        
        for g in genesplit:
            if g not in notes_dict:
                notes_dict[g]={}
            for s in snpsplit:
                if s not in notes_dict[g]:
                    notes_dict[g][s] = {}
                snp_notes = []
                for n in note.split('+'):
                    snp_notes.append(n.rstrip().lstrip())
                notes_dict[g][s] = snp_notes
        
    return notes_dict

def get_testi_auto(testi_auto_file):
     filein = testi_auto_file
     testi = pd.read_excel(filein)
     testi_dict = testi.set_index(testi.columns[0]).T.to_dict('list')
     return testi_dict

def cleansheet(sheet):
    for i, row in sheet.iterrows():
        if pd.isnull(row["GEN"]) or pd.isnull(row["GENE"]):
            sheet.drop(i, inplace=True)
    return sheet

# def cleanquery(sheet):


def build_scores_dicts(variants_list_file):
    filein = variants_list_file

    peso = pd.read_excel(filein, sheet_name='RIS AUMENTO DI PESO')
    peso = cleansheet(peso)

    #print(peso)
    scores_peso = calc_weights_dict(peso)


    t2d = pd.read_excel(filein, sheet_name='RIS T2D')
    t2d = cleansheet(t2d)
    #print(t2d)
    scores_t2d = calc_weights_dict(t2d)


    cardio = pd.read_excel(filein, sheet_name='RISC PATOLOGIE CARDIOVASCOLARI')
    cardio = cleansheet(cardio)
    #print(cardio)
    scores_cardio = calc_weights_dict(cardio)
    
    mamma = pd.read_excel(filein, sheet_name='RIS MAMMA')
    mamma = cleansheet(mamma)
    scores_mamma = calc_weights_dict(mamma)
    notes_mamma = get_notes_dict(mamma)
    # print(notes_mamma)


    plus = pd.read_excel(filein, sheet_name='RIS PLUS')
    plus = cleansheet(plus)
    scores_plus = calc_weights_dict(plus)
    notes_plus = get_notes_dict(plus)
    # print(scores_plus)
    # print(notes_plus)
    # print(notes_plus)


    vita = pd.read_excel(filein, sheet_name='RIS VITA')
    vita = cleansheet(vita)
    scores_vita = calc_weights_dict(vita)
    notes_vita = get_notes_dict(vita)
    # print(notes_vita)
    # print(scores_vita)

    sport = pd.read_excel(filein, sheet_name='RIS SPORT')
    sport = cleansheet(sport)
    scores_sport = calc_weights_dict(sport)
    notes_sport = get_notes_dict(sport)
    # print(scores_sport)
    # print(notes_sport)

    aging = pd.read_excel(filein, sheet_name='RIS AGEING')
    aging = cleansheet(aging)
    scores_aging = calc_weights_dict(aging)
    notes_aging = get_notes_dict(aging)
    # print(notes_aging)

    junior_intoll = pd.read_excel(filein, sheet_name='RIS JUNIOR_INTOLL')
    junior_intoll = cleansheet(junior_intoll)
    scores_junior_intoll = calc_weights_dict(junior_intoll)
    notes_junior_intoll = get_notes_dict(junior_intoll)
    # print(scores_junior_intoll)
    # print(notes_junior_intoll)

    junior_frag = pd.read_excel(filein, sheet_name='RIS JUNIOR_FRAG')
    junior_frag = cleansheet(junior_frag)
    scores_junior_frag = calc_weights_dict(junior_frag)
    notes_junior_frag = get_notes_dict(junior_frag)
    # print(scores_junior_frag)
    # print(notes_junior_frag)

    junior_met = pd.read_excel(filein, sheet_name='RIS JUNIOR_MET')
    junior_met = cleansheet(junior_met)
    scores_junior_met = calc_weights_dict(junior_met)
    notes_junior_met = get_notes_dict(junior_met)
    # print(scores_junior_met)
    # print(notes_junior_met)

    junior_carie = pd.read_excel(filein, sheet_name='RIS JUNIOR_CARIE')
    junior_carie = cleansheet(junior_carie)
    scores_junior_carie = calc_weights_dict(junior_carie)
    notes_junior_carie = get_notes_dict(junior_carie)
    # print(scores_junior_carie)
    # print(notes_junior_carie)


    return scores_peso, scores_t2d, scores_cardio, scores_mamma, notes_mamma, scores_plus, \
        notes_plus, scores_vita, notes_vita, scores_sport, notes_sport, scores_aging, notes_aging, \
        scores_junior_intoll, notes_junior_intoll, scores_junior_frag, notes_junior_frag, \
        scores_junior_met, notes_junior_met, scores_junior_carie, notes_junior_carie


def read_NGS_results_from_file(variants_file, packages, all_scores_notes, heet_name="Sheet1", as_dict=False,):
    print(packages)
    #Add "Condizioni" to the 'packages' list
    if "Base" in packages:
        packages.append("peso")
        packages.append("t2d")
        packages.append("cardio")
        packages.remove("Base")

    if "Junior" in packages:
        packages.append("Junior_sindrome_met")
        packages.append("Junior_intolleranze")
        packages.append("Junior_carie")
        packages.append("Junior_fragilita")
        packages.remove("Junior")
    print("variants_file:", variants_file)
    ngs_variants_df =build_pandas_variant_db(variants_file)
    for pkg in packages:
        if 'scores_'+pkg.lower() in all_scores_notes:
            print("Applying scores for package:", pkg)


    return ngs_variants_df



def read_query(metadata_file):      
    p = 'Condizioni'
    # Packages is an array containing one or more from ['Base', 'Plus', 'Vita', 'Sport', 'Ageing', 'Mamma']
    errors = []
    
    try:
        wb = load_workbook(filename = metadata_file)
        parsed_sheet = wb['Foglio '+p]
    except KeyError:
        errors.append("Errore: il foglio {0} non esiste nel file caricato".format(p))

    except Exception as exc:
        errors.append("Errore durante l'apertura del foglio {0}: {1}".format(p, exc))


    print("Processing sheet Foglio", p)
    # # print content of the sheet
    # for row in parsed_sheet.iter_rows(min_row=3, max_row=14, min_col=1, max_col=parsed_sheet.max_column+10):
    #     row_values = [cell.value for cell in row]
    #     print(row_values)

    try:
        rows_iter = parsed_sheet.iter_rows(min_col = 1, min_row = 3, max_col = parsed_sheet.max_column+10, max_row = 15)
        pazienti = [[cell.value for cell in list(row)[1:] if cell.value!=None] for row in rows_iter]
    except Exception as exc:
        errors.append("Errore durante la lettura dei dati nel foglio {0}: {1}".format(p, exc))
        

    print("PAZIENTI DATA EXTRACTED:")
    print(pazienti)

    if not pazienti:
        errors.append("Errore: nessun paziente trovato nel foglio {0}".format(p))
        

    # Assert that all pazienti lists have the same length
    legend = pazienti[0]  # Assuming the first row contains patient identifiers
    print("LEGEND:")
    print(legend)
    if not all(len(row) == len(legend) for row in pazienti):
        errors.append("Errore: qualcosa non va nel foglio Excel dei dati paziente. Controlla se qualcuno ha una cella vuota. ")
        


    # error = check_missing_data(pazienti, p, legend)
    # if error: # in case something is missing   
    #     errors.append(error)
    #     continue #do not proceed further with the sheet
    
    pazienti_dict = {}
    # print("Processing patients for sheet", p)
    # print("Number of patients:", len(pazienti[0]))
    # print(pazienti[0])
    for i in range(0, len(pazienti[0])):
        try:
            print(pazienti[0][i])
            patient_code = str(pazienti[0][i])
            print(patient_code)
            patient_id = str(pazienti[1][i])
            print(patient_id)
            patient_name = pazienti[2][i]
            print(patient_name)
            patient_email = pazienti[3][i]
            print(patient_email)
            patient_cf = pazienti[4][i]
            print(patient_cf)
            patient_weight = pazienti[5][i]
            print(patient_weight)
            patient_height = pazienti[6][i]
            print(patient_height)
            patient_lim_raw = pazienti[8][i]
            patient_lim = patient_lim_raw.lower() if patient_lim_raw else ''
            print(patient_lim)
            patient_dob = pazienti[9][i]
            print(patient_dob)
            patient_committent = pazienti[10][i]
            print(patient_committent)
            patient_condizioni = pazienti[11][i]
            print(patient_condizioni)
            patient_glutine = pazienti[12][i]
            print(patient_glutine)

            if len(patient_code) != 8:
                print("Wrong patient code for", patient_name, "code:", patient_code)
                errors.append(check_wrong_data(patient_name))
                continue
            
            
            if p=='Mamma':
                patient_gest = pazienti[7][i]
                pazienti_dict[patient_id] = {'code':patient_code, 'id':patient_id, 'name':patient_name, 'email':patient_email, 'cf':patient_cf, ':peso':patient_weight, 'altezza':patient_height, 'gestazione':patient_gest, 'lim':patient_lim, 'DOB':patient_dob,  'committent': patient_committent , 'condizioni': patient_condizioni, 'glutine': patient_glutine} 
            else:
                patient_sex = pazienti[7][i]
                pazienti_dict[patient_id] = {'code':patient_code, 'id':patient_id, 'name':patient_name, 'email':patient_email, 'cf':patient_cf, 'peso':patient_weight, 'altezza':patient_height, 'sesso':patient_sex, 'lim':patient_lim, 'DOB':patient_dob, 'committent': patient_committent, 'condizioni': patient_condizioni, 'glutine': patient_glutine} 
        except Exception as exc:
            errors.append("Errore durante l'elaborazione del paziente {0} nel foglio {1}: {2}".format(i+1, p, exc))
            continue
    
    # if p!='Condizioni':
    #     sheet =  pd.read_excel(variants_file, sheet_name='Foglio '+p, skiprows=13, index_col=0, names=["Gene", "SNP", "WT", "alt"]+list(pazienti_dict.keys()))
    # else:
    #     sheet =  pd.read_excel(metadata_file, sheet_name='Foglio '+p, skiprows=12, nrows=1)
    #     names=list(pazienti_dict.keys())
    #     print("NAMES IN CONDIZIONI", names)
    #     sheet.columns = names
    #     sheet = sheet.fillna(' ')
    #     print("SHEET IN CONDIZIONI")
    #     print(sheet)
    #     input("WAIT CONDIZIONI")

    pazienti_dict
    print("THIS IS ALLRESULTS FROMM XLSXREADER for", p)
    print(pazienti_dict)

    if errors:
        raise ValidationError(';'.join(errors))
    return pazienti_dict
